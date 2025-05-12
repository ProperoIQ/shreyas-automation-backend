from openpyxl import load_workbook
import os

def remove_decimals_from_excel(file_path):
    """
    Remove decimal places from all numeric cells in ALL sheets of an Excel file.
    """
    try:
        if not os.path.exists(file_path):
            print(f"⚠️ File not found: {file_path}")
            return

        print(f"🔧 Processing: {file_path}")
        wb = load_workbook(filename=file_path)
        
        for sheet in wb.worksheets:
            print(f"  📊 Processing sheet: {sheet.title}")
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        # Convert to integer (removes decimals)
                        cell.value = int(cell.value)
        
        wb.save(file_path)
        print(f"✅ Successfully processed: {file_path}")
    except Exception as e:
        print(f"❌ Error processing {file_path}: {str(e)}")
        raise