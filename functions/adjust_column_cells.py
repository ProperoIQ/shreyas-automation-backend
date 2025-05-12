import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def adjust_column_widths(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column  # 1-based index
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        sheet.column_dimensions[column_letter].width = max_length + 2

def convert_csv_to_xlsx_and_replace(csv_path):
    try:
        df = pd.read_csv(csv_path)
        temp_xlsx_path = csv_path.replace(".csv", "_temp.xlsx")
        
        # Save to temp xlsx
        df.to_excel(temp_xlsx_path, index=False)

        # Load, format column widths
        wb = load_workbook(temp_xlsx_path)
        sheet = wb.active
        adjust_column_widths(sheet)
        wb.save(temp_xlsx_path)

        # Replace CSV with formatted XLSX
        os.remove(csv_path)
        final_xlsx_path = csv_path.replace(".csv", ".xlsx")
        os.rename(temp_xlsx_path, final_xlsx_path)

        print(f"Converted and replaced: {os.path.basename(final_xlsx_path)}")
    except Exception as e:
        print(f"Failed to convert/replace {csv_path}: {e}")

      
def move_xlsx_to_output(csvdata_folder, output_folder):
    try:
        for filename in os.listdir(csvdata_folder):
            file_path = os.path.join(csvdata_folder, filename)

            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                # Copy the xlsx file to the 'output' folder
                dest_path = os.path.join(output_folder, filename)
                shutil.copy2(file_path, dest_path)
                print(f"Copied {filename} to {output_folder}")
    except Exception as e:
        print(f"Failed to copy files from {csvdata_folder} to {output_folder}: {e}")

def process_output_folder(folder_path="output"):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            try:
                wb = load_workbook(file_path)
                print(f"Adjusting column widths for: {filename}")
                for sheet in wb.worksheets:
                    adjust_column_widths(sheet)
                wb.save(file_path)
            except Exception as e:
                print(f"Failed to format {filename}: {e}")

        elif filename.endswith(".csv"):
            convert_csv_to_xlsx_and_replace(file_path)



