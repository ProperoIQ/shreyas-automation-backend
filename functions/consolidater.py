import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side

def process_and_merge_files(file1_path, file2_path, output_file_path):
    # Assuming file1_path is for NVB (input_customer_balance_nvb), file2_path for SMCS (input_customer_balance_smcs)
    # Read the Excel files into DataFrames
    file1 = pd.read_excel(file1_path, dtype=str)  # NVB
    file2 = pd.read_excel(file2_path, dtype=str)  # SMCS

    # Drop unnecessary columns (removed client data drops to keep them)
    columns_to_drop = ['customer_id', 'currency_id', 'contact']  # Adjusted to keep client data
    file1.drop(columns=[col for col in columns_to_drop if col in file1.columns], inplace=True)
    file2.drop(columns=[col for col in columns_to_drop if col in file2.columns], inplace=True)

    # Clean currency symbols for financial columns
    currency_columns = ["bcy_invoice_balance", "bcy_available_credits", "closing_balance"]
    for col in currency_columns:
        if col in file1.columns:
            file1[col] = pd.to_numeric(file1[col], errors='coerce')
        if col in file2.columns:
            file2[col] = pd.to_numeric(file2[col], errors='coerce')

    # Merge DataFrames
    unified_file = pd.merge(file2, file1, on="customer_name", how="outer", suffixes=("_file1", "_file2"))  # left=SMCS (_file1), right=NVB (_file2)

    # No longer dropping client data columns
    # columns_to_delete = [...]  # Commented out

    # Fill NaN with 0 for financials
    for col in currency_columns:
        if f"{col}_file1" in unified_file.columns:
            unified_file[f"{col}_file1"] = unified_file[f"{col}_file1"].fillna(0)
        if f"{col}_file2" in unified_file.columns:
            unified_file[f"{col}_file2"] = unified_file[f"{col}_file2"].fillna(0)

    # Add consolidated columns
    unified_file['Consolidated invoiced_amount'] = unified_file['bcy_invoice_balance_file1'] + unified_file['bcy_invoice_balance_file2']
    unified_file['Consolidated amount_received'] = unified_file['bcy_available_credits_file1'] + unified_file['bcy_available_credits_file2']
    unified_file['Consolidated closing_balance'] = unified_file['closing_balance_file1'] + unified_file['closing_balance_file2']

    # Define column order, including client data columns with suffixes
    desired_order = [
        "customer_name",
        "bcy_invoice_balance_file1", "bcy_available_credits_file1", "closing_balance_file1",
        "bcy_invoice_balance_file2", "bcy_available_credits_file2", "closing_balance_file2",
        "Consolidated invoiced_amount", "Consolidated amount_received", "Consolidated closing_balance",
        # SMCS client data (_file1 from SMCS)
        "last_name_file1", "email_file1", "mobile_phone_file1", 
        "contact.CF.Client Coordinator_file1", "contact.CF.Leadership_file1", 
        "contact.CF.Is Customer part of the Group of Companies_file1",
        # NVB client data (_file2 from NVB)
        "last_name_file2", "email_file2", "mobile_phone_file2", 
        "contact.CF.Client Coordinator_file2", "contact.CF.Leadership_file2", 
        "contact.CF.Is Customer part of the Group of Companies_file2"
    ]
    ordered_columns = [col for col in desired_order if col in unified_file.columns]
    unified_file = unified_file[ordered_columns]

    # Create MultiIndex column headers (top row labels)
    level1 = ["", 
              "SMCS receivables", "SMCS receivables", "SMCS receivables",
              "NVB receivables", "NVB receivables", "NVB receivables",
              "Consolidated", "Consolidated", "Consolidated",
              "SMCS Client Data", "SMCS Client Data", "SMCS Client Data", "SMCS Client Data", "SMCS Client Data", "SMCS Client Data",
              "NVB Client Data", "NVB Client Data", "NVB Client Data", "NVB Client Data", "NVB Client Data", "NVB Client Data"
             ]
    level2 = unified_file.columns.tolist()

    # Create Excel workbook and write data
    wb = Workbook()
    ws = wb.active

    # Define border style
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))

    # Write headers (two rows for MultiIndex)
    for col_num, (l1, l2) in enumerate(zip(level1, level2), 1):
        ws.cell(row=1, column=col_num, value=l1).border = thin_border
        ws.cell(row=2, column=col_num, value=l2).border = thin_border

    # Write data rows
    for r_idx, row in enumerate(dataframe_to_rows(unified_file, index=False, header=False), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    # Define colors
    smcs_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for SMCS
    nvb_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for NVB
    consolidated_color = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red for Consolidated
    client_data_color = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue for client data (optional)

    # Apply colors and borders to entire columns
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        header_value = col[0].value

        if header_value == "SMCS receivables" or header_value == "SMCS Client Data":
            fill = smcs_color
        elif header_value == "NVB receivables" or header_value == "NVB Client Data":
            fill = nvb_color
        elif header_value == "Consolidated":
            fill = consolidated_color
        else:
            fill = None

        for cell in col:
            if fill:
                cell.fill = fill
            cell.border = thin_border  # Ensure border is applied to all cells

    # Save the workbook
    excel_output_path = output_file_path.replace('.csv', '.xlsx')
    wb.save(excel_output_path)