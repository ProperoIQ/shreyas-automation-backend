import shutil
import os
import pandas as pd
import requests
from fastapi import FastAPI, Query, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from googleapiclient.errors import HttpError

from functions.remove_decimals import remove_decimals_from_excel
from functions.segregator import process_multiple_files
from functions.consolidater import process_and_merge_files
from functions.balance_summary import process_file
from functions.age_summary import generate_summary
from functions.combiner import combine_sheets
from functions.get_details import fetch_all_reports
from functions.get_invoices import invoice_step
from functions.adjust_column_cells import process_output_folder, move_xlsx_to_output

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://automation-frontend-589889616484.asia-south1.run.app",
        "http://localhost:3000"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def get_google_sheets_service():
    """
    Authenticate and return a Google Sheets API service client.
    """
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('sheets', 'v4', credentials=creds)

def fetch_and_save_google_sheet_json(spreadsheet_id: str, range_name: str, api_key: str, output_path: str):
    """
    Fetch data from Google Sheet using the Sheets API and save it as a JSON file.
    
    Args:
        spreadsheet_id (str): The ID of the Google Sheet.
        range_name (str): The range to fetch (e.g., 'Sheet1!A1:D').
        api_key (str): Google API key.
        output_path (str): Path to save the JSON file.
    
    Returns:
        dict: Status of the operation.
    """
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{range_name}?key={api_key}"
        response = requests.get(url)
        
        if response.status_code == 200:
            with open(output_path, "w") as f:
                f.write(response.text)
            print(f"[Google Sheet] JSON data saved to {output_path}")
            return {"status": "success", "message": f"JSON saved to {output_path}"}
        else:
            error_message = f"Failed to fetch Google Sheet data: {response.status_code} - {response.text}"
            print(f"[Google Sheet Error] {error_message}")
            return {"status": "error", "message": error_message}
    except Exception as e:
        error_message = f"Error fetching Google Sheet data: {str(e)}"
        print(f"[Google Sheet Error] {error_message}")
        return {"status": "error", "message": error_message}

def update_google_sheet_with_excel(spreadsheet_id: str, range_name: str, excel_path: str):
    """
    Update a Google Sheet with data from an Excel file.
    
    Args:
        spreadsheet_id (str): The ID of the Google Sheet.
        range_name (str): The range to update (e.g., 'Sheet1!A1').
        excel_path (str): Path to the Excel file.
    
    Returns:
        dict: Status of the operation.
    """
    try:
        # Read the Excel file
        excel_file = pd.ExcelFile(excel_path)
        service = get_google_sheets_service()
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            # Convert DataFrame to list of lists for Google Sheets API
            values = [df.columns.tolist()] + df.values.tolist()
            
            # Prepare the range for this sheet (create new sheet if needed)
            sheet_range = f"{sheet_name}!{range_name}"
            
            # Check if sheet exists, create if not
            sheets = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute().get('sheets', [])
            sheet_exists = any(sheet['properties']['title'] == sheet_name for sheet in sheets)
            if not sheet_exists:
                service.spreadsheets().sheets().create(
                    spreadsheetId=spreadsheet_id,
                    body={'properties': {'title': sheet_name}}
                ).execute()
            
            # Clear existing data in the range
            service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range=sheet_range
            ).execute()
            
            # Update the sheet with new data
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=sheet_range,
                valueInputOption="RAW",
                body={"values": values}
            ).execute()
            print(f"[Google Sheet] Updated sheet {sheet_name} in Google Sheet")
        
        return {"status": "success", "message": "Google Sheet updated successfully"}
    except Exception as e:
        error_message = f"Error updating Google Sheet: {str(e)}"
        print(f"[Google Sheet Error] {error_message}")
        return {"status": "error", "message": error_message}

def cleanup_folders(folders: list = ["csvdata", "output"], extra_files: list = ["output.zip"]):
    """
    Clean up specified folders by removing all files and subdirectories.
    Also removes specified extra files.
    """
    try:
        for folder in folders:
            if os.path.exists(folder):
                for item in os.listdir(folder):
                    item_path = os.path.join(folder, item)
                    try:
                        if os.path.isfile(item_path):
                            os.remove(item_path)
                            print(f"[Cleanup] Removed file: {item_path}")
                        elif os.path.isdir(item_path):
                            shutil.rmtree(item_path)
                    except Exception as e:
                        print(f"[Cleanup Error] Failed to process {item_path}: {str(e)}")
                print(f"[Cleanup] Successfully cleaned folder: {folder}")
            else:
                print(f"[Cleanup] Folder does not exist: {folder}")

        for file in extra_files:
            if os.path.exists(file):
                try:
                    os.remove(file)
                    print(f"[Cleanup] Removed file: {file}")
                except Exception as e:
                    print(f"[Cleanup Error] Failed to remove {file}: {str(e)}")
            else:
                print(f"[Cleanup] File does not exist: {file}")

        return {"status": "success", "message": "Cleanup completed successfully"}
    except Exception as e:
        return {"status": "error", "message": f"Cleanup failed: {str(e)}"}

def create_combined_excel(output_file: str, files_to_process: list):
    """
    Combine multiple Excel files into a single Excel file with separate sheets.
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for file_path in files_to_process:
            if os.path.exists(file_path):
                try:
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        df = excel_file.parse(sheet_name)
                        base_name = os.path.splitext(os.path.basename(file_path))[0]
                        if len(excel_file.sheet_names) > 1:
                            safe_sheet_name = f"{base_name[:15]}_{sheet_name[:15]}"[:31]
                        else:
                            safe_sheet_name = base_name[:31]
                        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                except Exception as e:
                    print(f"[Combine Error] {file_path}: {e}")

def add_hyperlinks(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    
    cons_sheet = None
    for sheet_name in wb.sheetnames:
        if 'consolidated' in sheet_name.lower():
            cons_sheet = wb[sheet_name]
            break
    
    if not cons_sheet:
        print("[Hyperlink] Consolidated sheet not found")
        wb.save(file_path)
        return
    
    smcs_aging_name = 'input_invoice_aging_smcs'
    nvb_aging_name = 'input_invoice_aging_nvb'
    
    if smcs_aging_name not in wb.sheetnames or nvb_aging_name not in wb.sheetnames:
        print("[Hyperlink] Aging sheets not found")
        wb.save(file_path)
        return
    
    smcs_aging_sheet = wb[smcs_aging_name]
    nvb_aging_sheet = wb[nvb_aging_name]
    
    def sort_aging_sheet(aging_sheet):
        cust_col = None
        for col in range(1, aging_sheet.max_column + 1):
            if aging_sheet.cell(1, col).value == 'customer_name':
                cust_col = col
                break
        if not cust_col:
            print(f"[Hyperlink] customer_name column not found in {aging_sheet.title}")
            return None
        
        data = []
        for r in range(2, aging_sheet.max_row + 1):
            row_data = [aging_sheet.cell(r, c).value for c in range(1, aging_sheet.max_column + 1)]
            data.append(row_data)
        
        data.sort(key=lambda x: x[cust_col - 1] if x[cust_col - 1] else '')
        
        for r in range(2, aging_sheet.max_row + 1):
            for c in range(1, aging_sheet.max_column + 1):
                aging_sheet.cell(r, c).value = None
        
        for i, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                aging_sheet.cell(i, c).value = val
        
        aging_sheet.auto_filter.ref = f"A1:{get_column_letter(aging_sheet.max_column)}{aging_sheet.max_row}"
        
        return cust_col
    
    smcs_cust_col = sort_aging_sheet(smcs_aging_sheet)
    nvb_cust_col = sort_aging_sheet(nvb_aging_sheet)
    
    if not smcs_cust_col or not nvb_cust_col:
        wb.save(file_path)
        return
    
    header_row = 2
    smcs_inv_col = None
    nvb_inv_col = None
    cust_name_col = None
    for col in range(1, cons_sheet.max_column + 1):
        cell_value = cons_sheet.cell(header_row, col).value
        if cell_value == 'Invoice Balance':
            header_group = cons_sheet.cell(1, col).value
            if header_group and 'SMCS Receivables' in header_group:
                smcs_inv_col = col
            elif header_group and 'NVB Receivables' in header_group:
                nvb_inv_col = col
        elif cell_value == 'customer_name':
            cust_name_col = col
    
    if not smcs_inv_col or not nvb_inv_col or not cust_name_col:
        print("[Hyperlink] Required columns not found in consolidated sheet")
        wb.save(file_path)
        return
    
    if 'Instructions' not in wb.sheetnames:
        inst_sheet = wb.create_sheet('Instructions')
        inst_sheet['A1'].value = "How to Use Hyperlinks"
        inst_sheet['A2'].value = (
            "Click a hyperlink to navigate to the aging sheet with a pre-applied filter for the exact company name. "
            "Use Excel's filter dropdown to adjust or clear the filter manually if needed."
        )
    
    for row in range(3, cons_sheet.max_row + 1):
        cust_name = cons_sheet.cell(row, cust_name_col).value
        if not cust_name:
            continue
        
        smcs_cell = cons_sheet.cell(row, smcs_inv_col)
        if isinstance(smcs_cell.value, (int, float)) and smcs_cell.value > 0:
            first_row = None
            for r in range(2, smcs_aging_sheet.max_row + 1):
                sheet_cust_name = smcs_aging_sheet.cell(r, smcs_cust_col).value
                if sheet_cust_name == cust_name:
                    first_row = r
                    break
            if first_row:
                smcs_cell.hyperlink = f"#'{smcs_aging_name}'!A{first_row}"
                smcs_cell.style = 'Hyperlink'
                smcs_cell.comment = openpyxl.comments.Comment(f"Filter for: {cust_name}", 'Grok')
                smcs_aging_sheet.auto_filter.add_filter_column(smcs_cust_col - 1, [cust_name])
        
        nvb_cell = cons_sheet.cell(row, nvb_inv_col)
        if isinstance(nvb_cell.value, (int, float)) and nvb_cell.value > 0:
            first_row = None
            for r in range(2, nvb_aging_sheet.max_row + 1):
                sheet_cust_name = nvb_aging_sheet.cell(r, nvb_cust_col).value
                if sheet_cust_name == cust_name:
                    first_row = r
                    break
            if first_row:
                nvb_cell.hyperlink = f"#'{nvb_aging_name}'!A{first_row}"
                nvb_cell.style = 'Hyperlink'
                nvb_cell.comment = openpyxl.comments.Comment(f"Filter for: {cust_name}", 'Grok')
                nvb_aging_sheet.auto_filter.add_filter_column(nvb_cust_col - 1, [cust_name])
    
    wb.save(file_path)
    print("[Hyperlink] Hyperlinks, auto-filters, and instructions added successfully")

def create_zip_archive(files_to_zip: list, zip_path: str):
    """
    Create a zip archive containing the specified files.
    """
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in files_to_zip:
                if os.path.exists(file_path):
                    zipf.write(file_path, os.path.basename(file_path))
                    print(f"[Zip] Added file to archive: {file_path}")
                else:
                    print(f"[Zip Error] File not found: {file_path}")
        print(f"[Zip] Created archive: {zip_path}")
    except Exception as e:
        print(f"[Zip Error] Failed to create zip archive: {str(e)}")
        raise

@app.post("/process_and_download")
async def process_and_download(
    background_tasks: BackgroundTasks,
    date_filter: str = Query(..., description="Date filter for fetching reports")
):
    try:
        # Step 0: Clean up previous files
        cleanup_folders()

        # Step 1: Fetch Google Sheet data and save as JSON
        spreadsheet_id = "129lcQiNSXYmV5SXHG54fjwsMfUkDCYFY9lYEZlKKJwo"
        range_name = "Sheet1!A1:D"
        api_key = "AIzaSyCd9e48q7cXwFwR0NaKEh_wJrEMn5nVyhM"
        json_output_path = "csvdata/google_sheet_data.json"
        
        os.makedirs("csvdata", exist_ok=True)
        os.makedirs("output", exist_ok=True)
        
        result = fetch_and_save_google_sheet_json(spreadsheet_id, range_name, api_key, json_output_path)
        if result["status"] == "error":
            raise HTTPException(status_code=500, detail=result["message"])

        # Step 2: Run processing pipeline
        fetch_all_reports(date_filter)
        process_multiple_files(
            'csvdata/input_invoice_aging_nvb.xlsx',
            'csvdata/input_invoice_aging_smcs.xlsx'
        )
        generate_summary(
            {
                'SMCS': 'output/SMCS_Age_Range_Columns.xlsx',
                'NVB': 'output/NVB_Age_Range_Columns.xlsx'
            },
            'output/Age_summary.xlsx'
        )
        process_and_merge_files(
            'csvdata/input_customer_balance_nvb.xlsx',
            'csvdata/input_customer_balance_smcs.xlsx',
            'output/unified_file.csv'
        )
        process_file('output/unified_file.xlsx', 'output/balances_summary.xlsx')
        combine_sheets('output/balances_summary.xlsx', 'output/Age_summary.xlsx', 'output/Final.xlsx')
       
        # Step 3: Collect all files (inputs + outputs)
        files_to_process = [
            'csvdata/input_invoice_aging_nvb.xlsx',
            'csvdata/input_invoice_aging_smcs.xlsx',
            'output/balances_summary.xlsx',
            'csvdata/input_customer_balance_nvb.xlsx',
            'csvdata/input_customer_balance_smcs.xlsx',
            'output/SMCS_Age_Range_Columns.xlsx',
            'output/NVB_Age_Range_Columns.xlsx',
            'output/Age_summary.xlsx',
            'output/unified_file.xlsx',
            'output/Final.xlsx',
            json_output_path
        ]

        # Step 4: Create a single combined Excel
        combined_file = "output/Combined_Report.xlsx"
        create_combined_excel(combined_file, files_to_process)
        
        # Step 5: Add hyperlinks to the consolidated sheet
        add_hyperlinks(combined_file)
        
        # Step 6: Update Google Sheet with Combined_Report.xlsx
        update_result = update_google_sheet_with_excel(spreadsheet_id, "A1", combined_file)
        if update_result["status"] == "error":
            raise HTTPException(status_code=500, detail=update_result["message"])
        
        files_to_process.append(combined_file)

        # Step 7: Create zip archive
        zip_path = "output.zip"
        create_zip_archive(files_to_process, zip_path)

        # Step 8: Return the combined Excel file as response
        return FileResponse(
            combined_file,
            filename="Combined_Report.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except FileNotFoundError as fnf_error:
        raise HTTPException(status_code=404, detail=f"File not found: {str(fnf_error)}")
    except PermissionError as perm_error:
        raise HTTPException(status_code=403, detail=f"Permission error: {str(perm_error)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")